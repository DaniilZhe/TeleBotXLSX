import openpyxl
import telebot
from openpyxl import workbook
from telebot import types

token = telebot.TeleBot('5567917612:AAFqPqeKOSVuDYWRepZVaZ5hS9aUKpACmCc')

#   TEXT VARIABLE

start_text = '''Здравствуйте.
Кнопки записывают текст в соответсвующие ячейки.  
После выбора поля - введите текст одним сообщением.
Будьте кратки и нформативны, либо нажмтите /help что бы увидеть пример заполнения. 
Что бы начать нажмите /go'''

help_text = '''Пример заполнения таблицы.

(A) Ситуация 
Муж повысил голос из-за того, что я долго собиралась на работу и ему пришлось меня ждать.

(B) Мысли
Он меня не любит.
Он меня ударит.

(С) Эмоция и её сила. «Картинка с эмоциями на заставке бота»
Обида – 80
Страх - 65

(D) Ощущения в теле
Ком в горле, напряжение лица. Боль в солнечном сплетении, напряжение в плечах.

(E) Поведение
Молча ускорилась.
Молчала всю дорогу до работы.
'''

#   CREATING TABLES

book: workbook = openpyxl.Workbook()
sheet = book.active

#   VARIABLE AND CELL NUMBER FOR TABLE

collum_num = 1
row_num = 2


#   COMMAND HELLO AND START

@token.message_handler(commands=['start'])
def start(command):
    token.reply_to(command, start_text)


@token.message_handler(commands=['help'])
def help(command):
    token.reply_to(command, help_text)


#   BUTTON

@token.message_handler(commands=['go'])
def button(go):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)
    chapter = types.KeyboardButton('Hовая запись')
    dowload_doc = types.KeyboardButton('Скачать')
    a = types.KeyboardButton('(A) Ситуация')
    b = types.KeyboardButton('(B) Мысли')
    c = types.KeyboardButton('(С) Эмоция и её сила')
    d = types.KeyboardButton('(D) Ощущения в теле')
    e = types.KeyboardButton('(E) Поведение')
    markup.add(a, b, c, d, e, chapter, dowload_doc)
    token.send_message(go.chat.id, 'Кнопки сформированы.\n<b>Нажмите кнопку «Ситауция»</b> и расскажите что произошло',
                       reply_markup=markup, parse_mode='html')


#   BUTTON CLICK TRACKING

@token.message_handler(content_types=['text'])
def tap_button(tap):
    global collum_num
    global row_num
    if tap.text == 'Hовая запись':
        token.register_next_step_handler(tap, user_mess)
        token.send_message(tap.chat.id, 'Теперь <b>не надо нажимать кнопку «Ситуация»</b> - вводите текст',
                           parse_mode='html')
        collum_num = 1
        row_num += 1
    if tap.text == '(A) Ситуация':
        token.register_next_step_handler(tap, user_mess)
    if tap.text == '(B) Мысли':
        token.register_next_step_handler(tap, user_mess)
        collum_num += 1
    if tap.text == '(С) Эмоция и её сила':
        token.register_next_step_handler(tap, user_mess)
        collum_num += 1
    if tap.text == '(D) Ощущения в теле':
        token.register_next_step_handler(tap, user_mess)
        collum_num += 1
    if tap.text == '(E) Поведение':
        token.register_next_step_handler(tap, user_mess)
        collum_num += 1
    if tap.text == 'Скачать':
        doc = open(f'Table_of_feelings.xlsx', 'rb')
        token.send_document(tap.chat.id, doc)


#   WRITING DATA FROM THE USER TO THE TABLE

def user_mess(ms):
    sheet.cell(column=1, row=1).value = 'Ситуация А'
    sheet.cell(column=2, row=1).value = 'Мысль В'
    sheet.cell(column=3, row=1).value = 'Эмоция и её сила С'
    sheet.cell(column=4, row=1).value = 'Ощущения в теле D'
    sheet.cell(column=5, row=1).value = 'Поведение Е'
    sheet.cell(column=collum_num, row=row_num).value = ms.text
    book.save(f'Table_of_feelings.xlsx')
    book.close()


token.polling(none_stop=True)
